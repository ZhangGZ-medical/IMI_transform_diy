"""
Slicer Python Console：导入deposit坐标到Markups Point List
在3D Slicer的Python Console (View → Python Interactor) 中执行。
"""

import os
from datetime import datetime


def generate_slicer_code(xlsx_path):
    """生成Slicer Python Console可执行的导入代码字符串"""
    fname = os.path.basename(xlsx_path)
    return f'''
# ===== Slicer Markups导入: deposit坐标 =====
# 源文件: {fname}
# 生成: IMI_transform_diy 技能

# 前置: 安装依赖(仅需一次)
# slicer.util.pip_install('openpyxl')

import openpyxl
import slicer

xlsx_path = r"{xlsx_path}"
wb = openpyxl.load_workbook(xlsx_path)

# 解析traj sheet
ws_traj = wb['traj']
traj_ids = set()
for row in ws_traj.iter_rows(min_row=1, values_only=True):
    if row[0] is None:
        continue
    label = str(row[0]).strip().lower()
    if label.startswith('e') or label.startswith('t'):
        traj_ids.add(label[1:])
traj_ids = sorted(traj_ids, key=int)
print(f"找到轨迹: {{traj_ids}}")

# 解析deposits sheet
ws_dep = wb['deposits']
rows = list(ws_dep.iter_rows(min_row=1, values_only=True))
header = rows[0]
num_deposits = sum(1 for v in header[1:] if v is not None)

r_row_idx = a_row_idx = s_row_idx = None
for i, row in enumerate(rows):
    label = str(row[0]).strip().lower() if row[0] is not None else ''
    if label == 'r': r_row_idx = i
    elif label == 'a': a_row_idx = i
    elif label == 's': s_row_idx = i

if None in (r_row_idx, a_row_idx, s_row_idx):
    raise ValueError("未找到R/A/S行")

default_traj_id = traj_ids[0]

# 创建point list并导入
point_list = slicer.mrmlScene.AddNewNodeByClass(
    'vtkMRMLMarkupsFiducialNode', f"traj{{default_traj_id}}_deposits")
point_list.SetMarkupLabelFormat('dep%-#5.0f')

for col_idx in range(1, num_deposits + 1):
    deposit_id = header[col_idx]
    r = float(rows[r_row_idx][col_idx])
    a = float(rows[a_row_idx][col_idx])
    s = float(rows[s_row_idx][col_idx])
    point_list.AddControlPoint([r, a, s], f"dep{{deposit_id}}")
    print(f"  dep{{deposit_id}}: R={{r:.4f}}, A={{a:.4f}}, S={{s:.4f}}")

print(f"\\n完成! Markups: traj{{default_traj_id}}_deposits ({{num_deposits}}点)")
'''


def save_slicer_md(xlsx_path, output_dir=None):
    """
    将生成的Slicer导入代码保存为带源文件名和时间戳的.md文档。
    文件名格式: slicer_import_{excel文件名去扩展名}_{年-月-日}_{时-分}.md
    例如: slicer_import_traj-15deposits_2026-06-02_09-30.md
    """
    code = generate_slicer_code(xlsx_path)
    stem = os.path.splitext(os.path.basename(xlsx_path))[0]
    ts = datetime.now().strftime('%Y-%m-%d_%H-%M')
    fname = f"slicer_import_{stem}_{ts}.md"

    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(xlsx_path))

    out_path = os.path.join(output_dir, fname)

    md_content = f"""# Slicer Python Console：导入deposit坐标到Markups Point List

> 源文件: `{xlsx_path}`
> 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
> 技能: IMI_transform_diy

## 前置：安装依赖

首先在Slicer Python Console中执行一次：

```python
slicer.util.pip_install('openpyxl')
```

安装完成后执行以下主代码：

```python
{code}
```
"""
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(md_content)

    print(f"Slicer导入代码已保存: {out_path}")
    return out_path
