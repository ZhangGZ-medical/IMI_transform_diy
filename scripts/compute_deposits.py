"""
导针入点→deposit RAS坐标计算器
几何模型：
  导针入点E(r1,a1,s1) → 导针尖端T(r2,a2,s2) 定义traj方向u
  自T沿u延伸depth → P_depth
  P_depth处开口(钟面角angle,12点=A轴投影=前方向;每钟点=30°;顺时针),细针与轴线夹角19°
  细针伸出exit → deposit位置

输出：回填deposits sheet的R/A/S列
"""
import math
import openpyxl
import sys

DEG = math.pi / 180.0
NEEDLE_ANGLE = 19.0  # 细针与导针轴线固定夹角(度)


def normalize(v):
    """归一化向量"""
    x, y, z = v
    mag = math.sqrt(x*x + y*y + z*z)
    if mag < 1e-12:
        return (0.0, 0.0, 0.0)
    return (x/mag, y/mag, z/mag)


def cross(a, b):
    """叉积"""
    return (a[1]*b[2] - a[2]*b[1],
            a[2]*b[0] - a[0]*b[2],
            a[0]*b[1] - a[1]*b[0])


def dot(a, b):
    """点积"""
    return a[0]*b[0] + a[1]*b[1] + a[2]*b[2]


def vec_add(a, b):
    return (a[0]+b[0], a[1]+b[1], a[2]+b[2])


def vec_scale(v, s):
    return (v[0]*s, v[1]*s, v[2]*s)


def compute_clock_direction(u_hat, angle_deg):
    """
    计算钟面方向(垂直于traj轴线的平面内)
    u_hat: 轨迹单位方向(entry→tip),即viewer look-at方向
    angle_deg: 顺时针角度,每钟点=30°(0°=12点=A轴投影,即前方向; 90°=3点)
    
    返回: 单位方向向量(在垂直于u_hat的平面内)
    """
    # 12点参考方向 = A轴(0,1,0)(前方向)在垂直平面上的投影
    a_axis = (0.0, 1.0, 0.0)
    a_dot_u = dot(a_axis, u_hat)
    
    # ref = a_axis - (a_axis·u_hat)*u_hat
    ref = vec_add(a_axis, vec_scale(u_hat, -a_dot_u))
    ref_mag = math.sqrt(dot(ref, ref))
    
    if ref_mag < 1e-10:
        # traj平行于A轴,改用S轴(0,0,1)作为参考
        s_axis = (0.0, 0.0, 1.0)
        s_dot_u = dot(s_axis, u_hat)
        ref = vec_add(s_axis, vec_scale(u_hat, -s_dot_u))
        ref_mag = math.sqrt(dot(ref, ref))
    
    ref_hat = vec_scale(ref, 1.0/ref_mag)
    
    # 垂直于ref_hat和u_hat的方向(右手定则: u_hat×ref_hat)
    perp = cross(u_hat, ref_hat)
    
    # 顺时针旋转: dir = cos(θ)*ref_hat - sin(θ)*(u_hat×ref_hat)
    # (viewer沿u_hat方向看,法向量指向viewer)
    theta = angle_deg * DEG
    cos_t = math.cos(theta)
    sin_t = math.sin(theta)
    
    dir_x = cos_t * ref_hat[0] - sin_t * perp[0]
    dir_y = cos_t * ref_hat[1] - sin_t * perp[1]
    dir_z = cos_t * ref_hat[2] - sin_t * perp[2]
    
    return normalize((dir_x, dir_y, dir_z))


def compute_deposit(entry, tip, depth, angle_deg, exit_len):
    """
    计算单个deposit的RAS坐标
    
    参数:
      entry: (r, a, s) 导针入点
      tip:   (r, a, s) 导针尖端
      depth: mm, 自尖端沿traj方向延伸距离
      angle_deg: 度, 钟面角(顺时针,0°=12点)
      exit_len: mm, 细针伸出长度
      
    返回:
      (r, a, s) deposit的RAS坐标
    """
    # 轨迹方向
    traj_vec = (tip[0]-entry[0], tip[1]-entry[1], tip[2]-entry[2])
    u_hat = normalize(traj_vec)
    
    # P_depth = tip + depth * u_hat
    p_depth = vec_add(tip, vec_scale(u_hat, depth))
    
    # 钟面方向(垂直平面内)
    clock_dir = compute_clock_direction(u_hat, angle_deg)
    
    # 细针方向: cos(19°)*u_hat + sin(19°)*clock_dir
    needle_cos = math.cos(NEEDLE_ANGLE * DEG)
    needle_sin = math.sin(NEEDLE_ANGLE * DEG)
    needle_dir = normalize(vec_add(
        vec_scale(u_hat, needle_cos),
        vec_scale(clock_dir, needle_sin)
    ))
    
    # deposit = P_depth + exit * needle_dir
    deposit = vec_add(p_depth, vec_scale(needle_dir, exit_len))
    
    return deposit


def parse_traj_sheet(ws):
    """
    解析traj sheet,返回 {traj_id: {entry: (r,a,s), tip: (r,a,s)}}
    格式: 每行(label, axis, value), 如 ('e1','r1',33.9)
    """
    trajs = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] is None:
            continue
        label = str(row[0]).strip().lower()  # e1, t1, e2, t2...
        axis = str(row[1]).strip().lower()    # r1, a1, s1, r2, a2, s2...
        value = float(row[2])
        
        # 提取traj编号: e1→1, t2→2
        if label.startswith('e'):
            traj_id = label[1:]  # '1', '2', ...
            if traj_id not in trajs:
                trajs[traj_id] = {'entry': [None, None, None], 'tip': [None, None, None]}
            idx = {'r': 0, 'a': 1, 's': 2}[axis[0]]
            trajs[traj_id]['entry'][idx] = value
        elif label.startswith('t'):
            traj_id = label[1:]
            if traj_id not in trajs:
                trajs[traj_id] = {'entry': [None, None, None], 'tip': [None, None, None]}
            idx = {'r': 0, 'a': 1, 's': 2}[axis[0]]
            trajs[traj_id]['tip'][idx] = value
    
    # 转换为tuple
    result = {}
    for tid, data in trajs.items():
        result[tid] = {
            'entry': tuple(data['entry']),
            'tip': tuple(data['tip'])
        }
    return result


def parse_deposits_sheet(ws):
    """
    解析deposits sheet,返回列式数据
    假设第1行为deposit#/traj标识,后续行为depth/angle/exit/R/A/S
    返回: list of dicts: [{deposit_id, traj_id, depth, angle, exit, row_idx_R, row_idx_A, row_idx_S}, ...]
    """
    # 读取所有数据
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []
    
    # 找出标题行
    header_row = None
    header_idx = -1
    for i, row in enumerate(rows):
        if row[0] is not None:
            first_cell = str(row[0]).strip().lower()
            if 'deposit' in first_cell or first_cell.startswith('traj'):
                header_row = row
                header_idx = i
                break
    
    if header_row is None:
        # 假设第1行是标题
        header_idx = 0
        header_row = rows[0]
    
    # 统计deposit列数(跳过第一列标签)
    num_deposits = 0
    for val in header_row[1:]:
        if val is not None:
            num_deposits += 1
        else:
            break
    
    # 建立行索引映射
    row_map = {}
    for i, row in enumerate(rows):
        if row[0] is not None:
            key = str(row[0]).strip().lower()
            row_map[key] = i + 1  # 1-based Excel行号
    
    deposits = []
    for col_idx in range(1, num_deposits + 1):
        deposit_id = str(header_row[col_idx]) if header_row[col_idx] is not None else str(col_idx)
        
        # 获取各参数
        depth = rows[row_map.get('depth', 1) - 1][col_idx] if 'depth' in row_map else None
        angle = rows[row_map.get('angle', 2) - 1][col_idx] if 'angle' in row_map else None
        exit_len = rows[row_map.get('exit', 3) - 1][col_idx] if 'exit' in row_map else None
        
        deposits.append({
            'deposit_id': deposit_id,
            'depth': float(depth) if depth is not None else 0.0,
            'angle': float(angle) if angle is not None else 0.0,
            'exit': float(exit_len) if exit_len is not None else 0.0,
            'col_idx': col_idx + 1,  # 1-based Excel列号
            'row_R': row_map.get('r', -1),
            'row_A': row_map.get('a', -1),
            'row_S': row_map.get('s', -1),
        })
    
    return deposits, row_map


def process_file(filepath):
    """处理单个Excel文件"""
    print(f"处理文件: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    
    # 解析traj
    ws_traj = wb['traj']
    trajs = parse_traj_sheet(ws_traj)
    print(f"  轨迹: {list(trajs.keys())}")
    for tid, data in trajs.items():
        print(f"    traj{tid}: E{data['entry']} → T{data['tip']}")
    
    # 解析deposits
    ws_dep = wb['deposits']
    deposits_info, row_map = parse_deposits_sheet(ws_dep)
    print(f"  Deposits: {len(deposits_info)}个")
    
    # 选择traj: 如果有多个traj且deposit_id可映射到traj_id则匹配,否则用第一个
    default_traj_id = list(trajs.keys())[0] if trajs else '1'
    
    # 计算每个deposit坐标
    results = []
    for dep in deposits_info:
        # 尝试匹配traj
        traj_id = str(dep['deposit_id'])
        if traj_id not in trajs:
            traj_id = default_traj_id
        
        traj_data = trajs.get(traj_id, trajs[default_traj_id])
        
        deposit_ras = compute_deposit(
            traj_data['entry'],
            traj_data['tip'],
            dep['depth'],
            dep['angle'],
            dep['exit']
        )
        
        results.append({
            **dep,
            'traj_id': traj_id,
            'R': round(deposit_ras[0], 4),
            'A': round(deposit_ras[1], 4),
            'S': round(deposit_ras[2], 4),
        })
        
        print(f"  deposit{dep['deposit_id']}: depth={dep['depth']} angle={dep['angle']}° exit={dep['exit']} → "
              f"R={results[-1]['R']:.4f} A={results[-1]['A']:.4f} S={results[-1]['S']:.4f}")
    
    # 回填R/A/S到deposits sheet
    # 查找R/A/S行: 如果已有则填入,否则追加
    if row_map.get('r', -1) > 0:
        r_row = row_map['r']
        a_row = row_map['a']
        s_row = row_map['s']
    else:
        # 需要找到或创建R/A/S行
        # 扫描现有行,找空行或最后一行
        max_row = ws_dep.max_row
        # 在最后添加R/A/S行
        r_row = max_row + 1
        a_row = max_row + 2
        s_row = max_row + 3
        ws_dep.cell(row=r_row, column=1, value='R')
        ws_dep.cell(row=a_row, column=1, value='A')
        ws_dep.cell(row=s_row, column=1, value='S')
    
    # 填入计算结果
    for res in results:
        col = res['col_idx']
        ws_dep.cell(row=r_row, column=col, value=res['R'])
        ws_dep.cell(row=a_row, column=col, value=res['A'])
        ws_dep.cell(row=s_row, column=col, value=res['S'])
    
    # 保存
    wb.save(filepath)
    print(f"  [OK] 已保存到 {filepath}")
    return results


if __name__ == '__main__':
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
    else:
        filepath = r'G:\phase2_ns_plan\coordinates transform\traj-15deposits.xlsx'
    
    process_file(filepath)
